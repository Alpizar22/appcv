"""
Analizador de CVs con IA / AI CV Analyzer — app.py
Streamlit app para reclutadores usando Claude (Anthropic)
"""
import io
import json
import os
import urllib.request
from typing import Optional

import anthropic
import pdfplumber
import streamlit as st
from dotenv import load_dotenv
from PIL import Image, ImageDraw, ImageFont

load_dotenv()

MODEL = "claude-sonnet-4-20250514"
GUMROAD_URL = "https://yourname.gumroad.com/l/your-product"  # ← actualiza esto / update this
MAX_USES = 30

APP_DIR    = os.path.dirname(os.path.abspath(__file__))
KEYS_FILE  = os.path.join(APP_DIR, "keys.txt")
USAGE_FILE = os.path.join(APP_DIR, "usage.json")
FONT_REG   = os.path.join(APP_DIR, "DejaVuSans.ttf")
FONT_BOLD  = os.path.join(APP_DIR, "DejaVuSans-Bold.ttf")


# ── Font download ──────────────────────────────────────────────────────────────

def ensure_fonts() -> bool:
    """Download DejaVu fonts if not present. Returns True if fonts are available."""
    base = "https://github.com/dejavu-fonts/dejavu-fonts/raw/master/ttf/"
    ok = True
    for path, name in [(FONT_REG, "DejaVuSans.ttf"), (FONT_BOLD, "DejaVuSans-Bold.ttf")]:
        if not os.path.exists(path):
            try:
                urllib.request.urlretrieve(base + name, path)
            except Exception:
                ok = False
    return ok


# ── Translations ───────────────────────────────────────────────────────────────

TEXTS: dict[str, dict] = {
    "es": {
        # Sidebar
        "lang_toggle": "EN",
        # Auth
        "auth_title": "🔐 Acceso requerido",
        "auth_subtitle": "Ingresa tu clave de acceso para continuar.",
        "auth_label": "Clave de acceso",
        "auth_placeholder": "Ej: CLAVE-XXXX-YYYY",
        "auth_btn": "Entrar",
        "auth_error_empty": "Ingresa una clave.",
        "auth_error_invalid": "Clave inválida. Verifica e intenta de nuevo.",
        "auth_error_expired": (
            f"Clave expirada — adquiere una nueva en "
            f"[este enlace]({GUMROAD_URL})"
        ),
        # Main UI
        "page_title": "Analizador de CVs con IA",
        "page_subtitle": "*Análisis inteligente para reclutadores — powered by Claude (Anthropic)*",
        "upload_header": "1 · Cargar CV",
        "upload_label": "Selecciona un CV en PDF",
        "upload_help": "El PDF debe tener texto seleccionable (no escaneado).",
        "upload_caption": "🔒 Este análisis se procesa en tiempo real y no almacenamos ningún dato del CV.",
        "jd_header": "2 · Descripción del puesto *(opcional)*",
        "jd_label": "Pega aquí la descripción del rol",
        "jd_placeholder": (
            "Ejemplo: Buscamos un Data Engineer con 3+ años en Python, Spark y AWS. "
            "Valoramos experiencia con pipelines ELT y conocimiento de dbt..."
        ),
        "analyze_btn": "🔍 Analizar CV",
        "analyze_caption": "Sube un PDF para habilitar el análisis.",
        # Spinners
        "spinner_extract": "Extrayendo texto e imagen del PDF…",
        "spinner_analyze": "Analizando con Claude…",
        "spinner_reports": "Generando reportes descargables…",
        # Errors
        "err_no_text": "No se pudo extraer texto. Verifica que el PDF no sea una imagen escaneada.",
        "err_no_api_key": "Falta ANTHROPIC_API_KEY en el archivo .env",
        "err_invalid_api": "API key inválida. Verifica tu ANTHROPIC_API_KEY.",
        "err_rate_limit": "Límite de tasa alcanzado. Espera e intenta de nuevo.",
        "err_api": "Error de la API: {}",
        "err_bad_json": "La IA devolvió un JSON inválido.",
        "err_raw_label": "Respuesta cruda",
        # Results
        "no_results": "👈 Sube un CV para ver el análisis aquí",
        "score_label": "Score General",
        "rec_label": "Recomendación",
        "verdict_header": "### 🎯 Veredicto Ejecutivo",
        "verdict_na": "No disponible.",
        "skills_header": "### 🛠️ Habilidades",
        "tab_tech": "Técnicas",
        "tab_soft": "Blandas",
        "tab_lang": "Idiomas",
        "no_tech": "No identificadas.",
        "no_soft": "No identificadas.",
        "no_lang": "No identificados.",
        "strong_header": "### ⭐ Puntos Fuertes",
        "strong_none": "Ninguno identificado.",
        "flags_header": "### 🚩 Red Flags",
        "flags_none": "Sin red flags detectadas.",
        "download_header": "### 📥 Descargar Resultados",
        "dl_pdf": "📄 Descargar PDF",
        "dl_excel": "📊 Descargar Excel",
        "json_expander": "Ver JSON completo",
        # Recommendation display
        "CONTRATAR": "CONTRATAR",
        "ENTREVISTAR": "ENTREVISTAR",
        "DESCARTAR": "DESCARTAR",
        # PDF strings
        "pdf_report_title": "Reporte de Análisis de CV",
        "pdf_label_email": "Email",
        "pdf_label_phone": "Teléfono",
        "pdf_score": "Score General",
        "pdf_verdict": "Veredicto Ejecutivo",
        "pdf_skills": "Habilidades",
        "pdf_tech": "Técnicas:",
        "pdf_soft": "Blandas:",
        "pdf_lang": "Idiomas:",
        "pdf_strong": "Puntos Fuertes",
        "pdf_flags": "Red Flags",
        "pdf_footer": (
            "Generado por Analizador de CVs con IA  |  Powered by Claude (Anthropic)  |  "
            "Este análisis es orientativo y no reemplaza el criterio del reclutador."
        ),
        # Excel strings
        "xl_sheet": "Análisis CV",
        "xl_title": "Reporte de Análisis de CV — Generado con IA",
        "xl_nombre": "Nombre",
        "xl_email": "Email",
        "xl_telefono": "Teléfono",
        "xl_linkedin": "LinkedIn",
        "xl_score": "Score General",
        "xl_rec": "Recomendación",
        "xl_verdict": "Veredicto",
        "xl_skill_type": "Tipo",
        "xl_skill_val": "Habilidades",
        "xl_tech": "Técnicas",
        "xl_soft": "Blandas",
        "xl_lang": "Idiomas",
        "xl_strong": "Puntos Fuertes",
        "xl_flags": "Red Flags",
        # ── Ranking mode ──
        "mode_single": "📄 Análisis Individual",
        "mode_ranking": "🏆 Ranking de Candidatos",
        "rank_jd_label": "Descripción del puesto *(obligatoria)*",
        "rank_jd_placeholder": "Describe el rol para que el ranking sea preciso. Ej: Buscamos un Senior Backend Engineer con experiencia en Python y microservicios...",
        "rank_upload_label": "CVs de candidatos (2 – 10 PDFs)",
        "rank_upload_help": "Sube hasta 10 PDFs con texto seleccionable. Cada CV cuenta como 1 uso.",
        "rank_caption": "🔒 Procesamiento en tiempo real. No almacenamos ningún dato.",
        "rank_btn": "🏆 Analizar y Rankear",
        "rank_btn_need_jd": "Escribe la descripción del puesto para habilitar el ranking.",
        "rank_btn_need_files": "Sube al menos 2 CVs para habilitar el ranking.",
        "rank_spinner": "Analizando CV {i}/{n}: {name}…",
        "rank_warn_failed": "⚠ {n} CV(s) no pudieron analizarse y se omitieron.",
        "rank_err_min": "Se necesitan al menos 2 resultados válidos para generar un ranking.",
        "rank_results_header": "## 🏆 Ranking de Candidatos",
        "rank_table_pos": "Pos.",
        "rank_table_name": "Candidato",
        "rank_table_score": "Score",
        "rank_table_rec": "Recomendación",
        "rank_table_skills": "Habilidades clave",
        "rank_table_verdict": "Resumen ejecutivo",
        "rank_detail_header": "### Detalle de candidatos",
        "rank_expand": "Ver detalle completo",
        "rank_download": "📊 Descargar Ranking Excel",
        "rank_no_results": "👆 Sube los CVs y describe el puesto para ver el ranking aquí",
        # Ranking Excel
        "xl_rank_sheet1": "Ranking",
        "xl_rank_sheet2": "Detalle",
        "xl_rank_title": "Ranking de Candidatos — Análisis con IA",
        "xl_rank_job": "Puesto analizado",
        "xl_rank_pos": "Posición",
        "xl_rank_file": "Archivo",
        "xl_rank_skills": "Habilidades técnicas",
    },
    "en": {
        # Sidebar
        "lang_toggle": "ES",
        # Auth
        "auth_title": "🔐 Access required",
        "auth_subtitle": "Enter your access key to continue.",
        "auth_label": "Access key",
        "auth_placeholder": "E.g.: KEY-XXXX-YYYY",
        "auth_btn": "Enter",
        "auth_error_empty": "Please enter a key.",
        "auth_error_invalid": "Invalid key. Please check and try again.",
        "auth_error_expired": (
            f"Key expired — get a new one at "
            f"[this link]({GUMROAD_URL})"
        ),
        # Main UI
        "page_title": "AI CV Analyzer",
        "page_subtitle": "*Intelligent analysis for recruiters — powered by Claude (Anthropic)*",
        "upload_header": "1 · Upload CV",
        "upload_label": "Select a PDF CV",
        "upload_help": "The PDF must have selectable text (not a scanned image).",
        "upload_caption": "🔒 This analysis is processed in real time and we do not store any CV data.",
        "jd_header": "2 · Job description *(optional)*",
        "jd_label": "Paste the job description here",
        "jd_placeholder": (
            "Example: We are looking for a Data Engineer with 3+ years in Python, Spark and AWS. "
            "We value experience with ELT pipelines and knowledge of dbt..."
        ),
        "analyze_btn": "🔍 Analyze CV",
        "analyze_caption": "Upload a PDF to enable analysis.",
        # Spinners
        "spinner_extract": "Extracting text and image from PDF…",
        "spinner_analyze": "Analyzing with Claude…",
        "spinner_reports": "Generating downloadable reports…",
        # Errors
        "err_no_text": "Could not extract text. Make sure the PDF is not a scanned image.",
        "err_no_api_key": "Missing ANTHROPIC_API_KEY in .env file",
        "err_invalid_api": "Invalid API key. Check your ANTHROPIC_API_KEY.",
        "err_rate_limit": "Rate limit reached. Wait and try again.",
        "err_api": "API error: {}",
        "err_bad_json": "The AI returned invalid JSON.",
        "err_raw_label": "Raw response",
        # Results
        "no_results": "👈 Upload a CV to see the analysis here",
        "score_label": "Overall Score",
        "rec_label": "Recommendation",
        "verdict_header": "### 🎯 Executive Verdict",
        "verdict_na": "Not available.",
        "skills_header": "### 🛠️ Skills",
        "tab_tech": "Technical",
        "tab_soft": "Soft Skills",
        "tab_lang": "Languages",
        "no_tech": "None identified.",
        "no_soft": "None identified.",
        "no_lang": "None identified.",
        "strong_header": "### ⭐ Strong Points",
        "strong_none": "None identified.",
        "flags_header": "### 🚩 Red Flags",
        "flags_none": "No red flags detected.",
        "download_header": "### 📥 Download Results",
        "dl_pdf": "📄 Download PDF",
        "dl_excel": "📊 Download Excel",
        "json_expander": "View full JSON",
        # Recommendation display
        "CONTRATAR": "HIRE",
        "ENTREVISTAR": "INTERVIEW",
        "DESCARTAR": "DISCARD",
        # PDF strings
        "pdf_report_title": "CV Analysis Report",
        "pdf_label_email": "Email",
        "pdf_label_phone": "Phone",
        "pdf_score": "Overall Score",
        "pdf_verdict": "Executive Verdict",
        "pdf_skills": "Skills",
        "pdf_tech": "Technical:",
        "pdf_soft": "Soft Skills:",
        "pdf_lang": "Languages:",
        "pdf_strong": "Strong Points",
        "pdf_flags": "Red Flags",
        "pdf_footer": (
            "Generated by AI CV Analyzer  |  Powered by Claude (Anthropic)  |  "
            "This analysis is indicative and does not replace recruiter judgement."
        ),
        # Excel strings
        "xl_sheet": "CV Analysis",
        "xl_title": "CV Analysis Report — AI Generated",
        "xl_nombre": "Name",
        "xl_email": "Email",
        "xl_telefono": "Phone",
        "xl_linkedin": "LinkedIn",
        "xl_score": "Overall Score",
        "xl_rec": "Recommendation",
        "xl_verdict": "Verdict",
        "xl_skill_type": "Type",
        "xl_skill_val": "Skills",
        "xl_tech": "Technical",
        "xl_soft": "Soft Skills",
        "xl_lang": "Languages",
        "xl_strong": "Strong Points",
        "xl_flags": "Red Flags",
        # ── Ranking mode ──
        "mode_single": "📄 Individual Analysis",
        "mode_ranking": "🏆 Candidate Ranking",
        "rank_jd_label": "Job description *(required)*",
        "rank_jd_placeholder": "Describe the role so the ranking is accurate. E.g.: We are looking for a Senior Backend Engineer with Python and microservices experience...",
        "rank_upload_label": "Candidate CVs (2 – 10 PDFs)",
        "rank_upload_help": "Upload up to 10 PDFs with selectable text. Each CV counts as 1 use.",
        "rank_caption": "🔒 Real-time processing. We do not store any data.",
        "rank_btn": "🏆 Analyze & Rank",
        "rank_btn_need_jd": "Write the job description to enable ranking.",
        "rank_btn_need_files": "Upload at least 2 CVs to enable ranking.",
        "rank_spinner": "Analyzing CV {i}/{n}: {name}…",
        "rank_warn_failed": "⚠ {n} CV(s) could not be analyzed and were skipped.",
        "rank_err_min": "At least 2 valid results are needed to generate a ranking.",
        "rank_results_header": "## 🏆 Candidate Ranking",
        "rank_table_pos": "Pos.",
        "rank_table_name": "Candidate",
        "rank_table_score": "Score",
        "rank_table_rec": "Recommendation",
        "rank_table_skills": "Key skills",
        "rank_table_verdict": "Executive summary",
        "rank_detail_header": "### Candidate details",
        "rank_expand": "View full detail",
        "rank_download": "📊 Download Ranking Excel",
        "rank_no_results": "👆 Upload CVs and describe the role to see the ranking here",
        # Ranking Excel
        "xl_rank_sheet1": "Ranking",
        "xl_rank_sheet2": "Detail",
        "xl_rank_title": "Candidate Ranking — AI Analysis",
        "xl_rank_job": "Analyzed role",
        "xl_rank_pos": "Position",
        "xl_rank_file": "File",
        "xl_rank_skills": "Technical skills",
    },
}

# ── System prompts ─────────────────────────────────────────────────────────────

SYSTEM_PROMPT = {
    "es": """Eres un experto reclutador técnico con más de 15 años de experiencia evaluando CVs.
Analiza el CV proporcionado y devuelve ÚNICAMENTE un objeto JSON válido (sin markdown, sin texto adicional) con esta estructura exacta:

{
  "contacto": {
    "nombre": "<nombre completo del candidato, o null si no aparece>",
    "email": "<email del candidato, o null>",
    "telefono": "<teléfono del candidato, o null>",
    "linkedin": "<URL de LinkedIn, o null>"
  },
  "score_general": <entero del 0 al 100>,
  "habilidades": {
    "tecnicas": [<lista de habilidades técnicas>],
    "blandas": [<lista de habilidades blandas>],
    "idiomas": [<lista de idiomas con nivel, ej: "Inglés - C1">]
  },
  "puntos_fuertes": [<lista de aspectos positivos destacados>],
  "red_flags": [<lista de alertas o puntos negativos>],
  "veredicto_ejecutivo": "<párrafo de 2-3 oraciones con el resumen final EN ESPAÑOL>",
  "recomendacion": "<CONTRATAR | ENTREVISTAR | DESCARTAR>"
}

Responde en español. Si se incluye una descripción de puesto, ajusta el score y el análisis al encaje con ese rol específico.""",

    "en": """You are an expert technical recruiter with over 15 years of experience evaluating CVs.
Analyze the provided CV and return ONLY a valid JSON object (no markdown, no additional text) with this exact structure:

{
  "contacto": {
    "nombre": "<candidate's full name, or null if not present>",
    "email": "<candidate's email, or null>",
    "telefono": "<candidate's phone, or null>",
    "linkedin": "<LinkedIn URL, or null>"
  },
  "score_general": <integer from 0 to 100>,
  "habilidades": {
    "tecnicas": [<list of technical skills>],
    "blandas": [<list of soft skills>],
    "idiomas": [<list of languages with level, e.g.: "Spanish - Native">]
  },
  "puntos_fuertes": [<list of highlighted positive aspects>],
  "red_flags": [<list of concerns or negative points>],
  "veredicto_ejecutivo": "<2-3 sentence paragraph with the final summary IN ENGLISH>",
  "recomendacion": "<CONTRATAR | ENTREVISTAR | DESCARTAR>"
}

Respond in English. If a job description is included, adjust the score and analysis to the fit with that specific role.""",
}


# ── Access key helpers ─────────────────────────────────────────────────────────

def _load_valid_keys() -> set:
    if not os.path.exists(KEYS_FILE):
        return set()
    with open(KEYS_FILE, "r", encoding="utf-8") as fh:
        return {line.strip() for line in fh if line.strip()}


def _load_usage() -> dict:
    if not os.path.exists(USAGE_FILE):
        return {}
    with open(USAGE_FILE, "r", encoding="utf-8") as fh:
        return json.load(fh)


def _save_usage(usage: dict) -> None:
    with open(USAGE_FILE, "w", encoding="utf-8") as fh:
        json.dump(usage, fh, indent=2)


def validate_key(key: str) -> str:
    """Return 'ok', 'invalid', or 'expired'."""
    if key not in _load_valid_keys():
        return "invalid"
    usage = _load_usage()
    if usage.get(key, 0) >= MAX_USES:
        return "expired"
    return "ok"


def consume_key(key: str) -> None:
    """Increment use count for the given key."""
    usage = _load_usage()
    usage[key] = usage.get(key, 0) + 1
    _save_usage(usage)


def key_uses_left(key: str) -> int:
    usage = _load_usage()
    return max(0, MAX_USES - usage.get(key, 0))


# ── CV helpers ────────────────────────────────────────────────────────────────

def extract_text_from_pdf(pdf_bytes: bytes, t: dict) -> str:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            return "\n".join(
                page.extract_text() or "" for page in pdf.pages
            ).strip()
    except Exception as e:
        st.error(t["err_no_text"] + f" ({e})")
        return ""


def extract_profile_image(pdf_bytes: bytes) -> Optional[Image.Image]:
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages[:2]:
                for img in page.images:
                    w = img.get("width") or img.get("Width") or 0
                    h = img.get("height") or img.get("Height") or 0
                    if not isinstance(w, (int, float)) or not isinstance(h, (int, float)):
                        continue
                    if w >= 60 and h >= 60:
                        stream = img.get("stream")
                        if stream is None:
                            continue
                        try:
                            data = stream.get_data()
                            return Image.open(io.BytesIO(data))
                        except Exception:
                            pass
    except Exception:
        pass
    return None


def make_initials_avatar(name: str) -> Image.Image:
    size = 100
    base = Image.new("RGB", (size, size), (91, 141, 239))
    draw = ImageDraw.Draw(base)
    words = [w for w in name.strip().split() if w]
    initials = "".join(w[0].upper() for w in words[:2]) or "?"
    try:
        font = ImageFont.load_default(size=38)
    except TypeError:
        font = ImageFont.load_default()
    try:
        bbox = draw.textbbox((0, 0), initials, font=font)
        tw, th = bbox[2] - bbox[0], bbox[3] - bbox[1]
    except AttributeError:
        tw, th = draw.textsize(initials, font=font)  # type: ignore[attr-defined]
    draw.text(((size - tw) / 2, (size - th) / 2 - 2), initials, fill="white", font=font)
    mask = Image.new("L", (size, size), 0)
    ImageDraw.Draw(mask).ellipse([0, 0, size - 1, size - 1], fill=255)
    result = Image.new("RGBA", (size, size), (0, 0, 0, 0))
    result.paste(base, mask=mask)
    return result


def analyze_cv(cv_text: str, job_description: str, lang: str, t: dict) -> Optional[dict]:
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        st.error(t["err_no_api_key"])
        return None

    client = anthropic.Anthropic(api_key=api_key)
    content = f"TEXTO DEL CV:\n{cv_text}"
    if job_description.strip():
        content += f"\n\n---\nJOB DESCRIPTION:\n{job_description.strip()}"

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=2048,
            system=SYSTEM_PROMPT[lang],
            messages=[{"role": "user", "content": content}],
        )
    except anthropic.AuthenticationError:
        st.error(t["err_invalid_api"])
        return None
    except anthropic.RateLimitError:
        st.error(t["err_rate_limit"])
        return None
    except anthropic.APIError as e:
        st.error(t["err_api"].format(e))
        return None

    raw = response.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.rsplit("```", 1)[0].strip()

    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        st.error(t["err_bad_json"])
        with st.expander(t["err_raw_label"]):
            st.code(raw)
        return None


# ── PDF report ────────────────────────────────────────────────────────────────

def generate_pdf_report(result: dict, t: dict) -> bytes:
    from fpdf import FPDF  # type: ignore

    ensure_fonts()
    use_unicode = os.path.exists(FONT_REG) and os.path.exists(FONT_BOLD)

    contacto  = result.get("contacto") or {}
    nombre    = contacto.get("nombre") or "Candidato"
    email     = contacto.get("email") or ""
    telefono  = contacto.get("telefono") or ""
    linkedin  = contacto.get("linkedin") or ""
    score     = int(result.get("score_general") or 0)
    rec_raw   = (result.get("recomendacion") or "").upper()
    rec_label = t.get(rec_raw, rec_raw)
    veredicto = result.get("veredicto_ejecutivo") or ""
    habs      = result.get("habilidades") or {}
    puntos    = result.get("puntos_fuertes") or []
    flags     = result.get("red_flags") or []

    REC_COLORS = {
        "CONTRATAR":   (40, 167, 69),
        "ENTREVISTAR": (255, 193, 7),
        "DESCARTAR":   (220, 53, 69),
    }
    rc = REC_COLORS.get(rec_raw, (108, 117, 125))
    sc = (40, 167, 69) if score >= 75 else (253, 126, 20) if score >= 50 else (220, 53, 69)

    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.add_page()
    pdf.set_left_margin(20)
    pdf.set_right_margin(20)

    # Register DejaVu TTF fonts if available (fpdf2 ≥2.7 detects unicode automatically)
    if use_unicode:
        try:
            pdf.add_font("DejaVu", "",  FONT_REG)
            pdf.add_font("DejaVu", "B", FONT_BOLD)
            FONT_NAME = "DejaVu"
        except Exception:
            use_unicode = False
            FONT_NAME = "Helvetica"
    else:
        FONT_NAME = "Helvetica"

    def _safe(text) -> str:
        """Sanitize text for PDF output. Always strips problematic unicode."""
        if not isinstance(text, str):
            text = str(text) if text is not None else ""
        # Replace common symbols that break cp1252 / core fonts
        text = (text
                .replace("\u2022", "-")   # •  bullet
                .replace("\u00b7", "-")   # ·  middle dot
                .replace("\u2192", "->")  # →  arrow
                .replace("\u2713", "OK")  # ✓  check mark
                .replace("\u2714", "OK")  # ✔  heavy check mark
                .replace("\u26a0", "!")   # ⚠  warning sign
                .replace("\u2605", "*")   # ★  star
                .replace("\u2019", "'")   # '  right single quote
                .replace("\u2018", "'")   # '  left single quote
                .replace("\u201c", '"')   # "  left double quote
                .replace("\u201d", '"')   # "  right double quote
                .replace("\u2013", "-")   # –  en dash
                .replace("\u2014", "-")   # —  em dash
                .replace("\u2026", "...")  # …  ellipsis
                )
        # Drop any remaining non-latin characters
        text = "".join(c if ord(c) < 256 else "?" for c in text)
        return text

    # ── Header ──
    pdf.set_fill_color(37, 99, 235)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(FONT_NAME, "B", 15)
    pdf.cell(0, 14, _safe(t["pdf_report_title"]), fill=True, align="C")
    pdf.ln(6)

    # ── Candidate info ──
    pdf.set_text_color(20, 20, 20)
    pdf.set_font(FONT_NAME, "B", 13)
    pdf.cell(0, 8, _safe(nombre))
    pdf.ln(8)

    pdf.set_font(FONT_NAME, "", 9)
    pdf.set_text_color(90, 90, 90)
    for label, val in [
        (t["pdf_label_email"], email),
        (t["pdf_label_phone"], telefono),
        ("LinkedIn", linkedin),
    ]:
        if val:
            pdf.cell(0, 5, _safe(f"{label}: {val}"))
            pdf.ln(5)
    pdf.ln(4)

    # ── Score ──
    pdf.set_text_color(20, 20, 20)
    pdf.set_font(FONT_NAME, "B", 11)
    pdf.cell(0, 6, _safe(f"{t['pdf_score']}: {score}/100"))
    pdf.ln(7)

    bar_x = pdf.l_margin
    bar_y = pdf.get_y()
    bar_w = pdf.w - pdf.l_margin - pdf.r_margin
    pdf.set_fill_color(220, 220, 220)
    pdf.rect(bar_x, bar_y, bar_w, 4, "F")
    filled = int(bar_w * score / 100)
    if filled > 0:
        pdf.set_fill_color(*sc)
        pdf.rect(bar_x, bar_y, filled, 4, "F")
    pdf.set_y(bar_y + 8)

    # ── Recommendation badge ──
    pdf.set_fill_color(*rc)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font(FONT_NAME, "B", 10)
    pdf.cell(55, 7, f"  {_safe(rec_label)}", fill=True)
    pdf.set_text_color(20, 20, 20)
    pdf.ln(11)

    # ── Section / bullet helpers ──
    def section(title: str):
        pdf.set_fill_color(243, 244, 246)
        pdf.set_font(FONT_NAME, "B", 11)
        pdf.cell(0, 7, _safe(title), fill=True)
        pdf.ln(8)

    def bullet(text: str, prefix: str = "-"):
        pdf.set_font(FONT_NAME, "", 10)
        pdf.multi_cell(0, 5, _safe(f"  {prefix} {text}"))
        pdf.ln(1)

    # ── Veredicto ──
    section(t["pdf_verdict"])
    pdf.set_font(FONT_NAME, "", 10)
    pdf.multi_cell(0, 5, _safe(veredicto))
    pdf.ln(5)

    # ── Habilidades ──
    section(t["pdf_skills"])
    for label, key in [
        (t["pdf_tech"], "tecnicas"),
        (t["pdf_soft"], "blandas"),
        (t["pdf_lang"], "idiomas"),
    ]:
        items = habs.get(key) or []
        if items:
            pdf.set_font(FONT_NAME, "B", 10)
            pdf.cell(0, 5, _safe(f"  {label}"))
            pdf.ln(5)
            for s in items:
                bullet(s)
            pdf.ln(2)
    pdf.ln(3)

    # ── Puntos fuertes ──
    if puntos:
        section(t["pdf_strong"])
        for p in puntos:
            bullet(p, "+")
        pdf.ln(4)

    # ── Red Flags ──
    if flags:
        section(t["pdf_flags"])
        for f in flags:
            bullet(f, "!")
        pdf.ln(4)

    # ── Footer ──
    pdf.set_font(FONT_NAME, "I" if not use_unicode else "", 8)
    pdf.set_text_color(150, 150, 150)
    pdf.multi_cell(0, 4, _safe(t["pdf_footer"]), align="C")

    return bytes(pdf.output())


# ── Excel report ──────────────────────────────────────────────────────────────

def generate_excel_report(result: dict, t: dict) -> bytes:
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore

    contacto = result.get("contacto") or {}
    habs     = result.get("habilidades") or {}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = t["xl_sheet"]

    BLUE   = "2563EB"
    LIGHT  = "F3F4F6"
    GREEN  = "D1FAE5"
    RED    = "FEE2E2"
    YELLOW = "FEF9C3"

    def hdr(cell, value):
        cell.value = value
        cell.font  = Font(bold=True, color="FFFFFF", size=11)
        cell.fill  = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    def lbl(cell, value):
        cell.value = value
        cell.font  = Font(bold=True, size=10)
        cell.fill  = PatternFill("solid", fgColor=LIGHT)
        cell.alignment = Alignment(vertical="center")

    def val(cell, value, fgColor=None):
        cell.value = value
        cell.font  = Font(size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if fgColor:
            cell.fill = PatternFill("solid", fgColor=fgColor)

    # ── Title ──
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = t["xl_title"]
    title_cell.font  = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill  = PatternFill("solid", fgColor=BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Contact + summary ──
    rec_raw   = result.get("recomendacion") or ""
    rec_label = t.get(rec_raw.upper(), rec_raw)
    fields = [
        (t["xl_nombre"],  contacto.get("nombre") or ""),
        (t["xl_email"],   contacto.get("email") or ""),
        (t["xl_telefono"],contacto.get("telefono") or ""),
        (t["xl_linkedin"],contacto.get("linkedin") or ""),
        (t["xl_score"],   f"{result.get('score_general', 0)}/100"),
        (t["xl_rec"],     rec_label),
        (t["xl_verdict"], result.get("veredicto_ejecutivo") or ""),
    ]
    REC_FILL = {"CONTRATAR": GREEN, "HIRE": GREEN,
                "ENTREVISTAR": YELLOW, "INTERVIEW": YELLOW,
                "DESCARTAR": RED, "DISCARD": RED}

    row = 2
    for label, value in fields:
        lbl(ws.cell(row=row, column=1), label)
        fill = REC_FILL.get(value.upper()) if label == t["xl_rec"] else None
        val(ws.cell(row=row, column=2), value, fill)
        ws.merge_cells(f"B{row}:C{row}")
        row += 1

    row += 1  # blank separator

    # ── Skills ──
    hdr(ws.cell(row=row, column=1), t["xl_skill_type"])
    hdr(ws.cell(row=row, column=2), t["xl_skill_val"])
    ws.merge_cells(f"B{row}:C{row}")
    row += 1

    for tipo, key in [
        (t["xl_tech"], "tecnicas"),
        (t["xl_soft"], "blandas"),
        (t["xl_lang"], "idiomas"),
    ]:
        items = habs.get(key) or []
        lbl(ws.cell(row=row, column=1), tipo)
        val(ws.cell(row=row, column=2), "\n".join(items) if items else "—")
        ws.merge_cells(f"B{row}:C{row}")
        ws.row_dimensions[row].height = max(15, len(items) * 14)
        row += 1

    row += 1

    # ── Puntos fuertes ──
    puntos = result.get("puntos_fuertes") or []
    if puntos:
        ws.merge_cells(f"A{row}:C{row}")
        hdr(ws.cell(row=row, column=1), t["xl_strong"])
        row += 1
        for p in puntos:
            ws.merge_cells(f"A{row}:C{row}")
            c = ws.cell(row=row, column=1, value=f"+ {p}")
            c.fill = PatternFill("solid", fgColor=GREEN)
            c.alignment = Alignment(wrap_text=True)
            row += 1
        row += 1

    # ── Red Flags ──
    flags = result.get("red_flags") or []
    if flags:
        ws.merge_cells(f"A{row}:C{row}")
        hdr(ws.cell(row=row, column=1), t["xl_flags"])
        row += 1
        for f in flags:
            ws.merge_cells(f"A{row}:C{row}")
            c = ws.cell(row=row, column=1, value=f"! {f}")
            c.fill = PatternFill("solid", fgColor=RED)
            c.alignment = Alignment(wrap_text=True)
            row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def analyze_cv_silent(cv_text: str, job_description: str, lang: str) -> tuple[Optional[dict], str]:
    """Like analyze_cv but returns (result, error_msg) without calling st.error."""
    api_key = os.getenv("ANTHROPIC_API_KEY")
    if not api_key:
        return None, "no_api_key"

    client = anthropic.Anthropic(api_key=api_key)
    content = f"TEXTO DEL CV:\n{cv_text}"
    if job_description.strip():
        content += f"\n\n---\nJOB DESCRIPTION:\n{job_description.strip()}"

    try:
        response = client.messages.create(
            model=MODEL,
            max_tokens=2048,
            system=SYSTEM_PROMPT[lang],
            messages=[{"role": "user", "content": content}],
        )
    except (anthropic.AuthenticationError, anthropic.RateLimitError, anthropic.APIError) as e:
        return None, str(e)

    raw = response.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.rsplit("```", 1)[0].strip()

    try:
        return json.loads(raw), ""
    except json.JSONDecodeError:
        return None, "json_error"


def generate_ranking_excel(ranked: list, job_description: str, t: dict) -> bytes:
    """
    ranked: list of {"rank": int, "filename": str, "result": dict}, sorted best-first.
    """
    import openpyxl  # type: ignore
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore

    BLUE   = "2563EB"
    LIGHT  = "F3F4F6"
    GREEN  = "D1FAE5"
    RED    = "FEE2E2"
    YELLOW = "FEF9C3"
    GOLD   = "FFF8DC"

    REC_FILL = {
        "CONTRATAR": GREEN, "HIRE": GREEN,
        "ENTREVISTAR": YELLOW, "INTERVIEW": YELLOW,
        "DESCARTAR": RED, "DISCARD": RED,
    }
    MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}

    def _hdr(cell, value, size=11):
        cell.value = value
        cell.font  = Font(bold=True, color="FFFFFF", size=size)
        cell.fill  = PatternFill("solid", fgColor=BLUE)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")

    def _lbl(cell, value):
        cell.value = value
        cell.font  = Font(bold=True, size=10)
        cell.fill  = PatternFill("solid", fgColor=LIGHT)
        cell.alignment = Alignment(vertical="center")

    def _val(cell, value, fg=None, bold=False):
        cell.value = value
        cell.font  = Font(size=10, bold=bold)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        if fg:
            cell.fill = PatternFill("solid", fgColor=fg)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ranking table ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = t["xl_rank_sheet1"]

    # Title row
    ws1.merge_cells("A1:G1")
    tc = ws1["A1"]
    tc.value = t["xl_rank_title"]
    tc.font  = Font(bold=True, size=14, color="FFFFFF")
    tc.fill  = PatternFill("solid", fgColor=BLUE)
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Job description row
    ws1.merge_cells("A2:G2")
    jd_short = job_description.strip()[:200] + ("…" if len(job_description.strip()) > 200 else "")
    ws1["A2"].value = f"{t['xl_rank_job']}: {jd_short}"
    ws1["A2"].font  = Font(italic=True, size=9, color="444444")
    ws1["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws1.row_dimensions[2].height = 20

    # Column headers (row 3)
    cols = [
        t["xl_rank_pos"], t["xl_nombre"], t["xl_email"],
        t["xl_score"],   t["xl_rec"],    t["xl_rank_skills"], t["rank_table_verdict"],
    ]
    for ci, h in enumerate(cols, 1):
        _hdr(ws1.cell(row=3, column=ci), h)
    ws1.row_dimensions[3].height = 20

    # Data rows
    for entry in ranked:
        rank    = entry["rank"]
        result  = entry["result"]
        fname   = entry["filename"]
        ctc     = result.get("contacto") or {}
        nombre  = ctc.get("nombre") or fname
        email   = ctc.get("email") or ""
        score   = int(result.get("score_general") or 0)
        rec_raw = (result.get("recomendacion") or "").upper()
        rec_dis = t.get(rec_raw, rec_raw)
        techs   = (result.get("habilidades") or {}).get("tecnicas") or []
        tech_s  = ", ".join(techs[:6]) if techs else "—"
        verdict = result.get("veredicto_ejecutivo") or ""
        medal   = MEDALS.get(rank, str(rank))
        row_fg  = GOLD if rank == 1 else None
        dr      = 3 + rank

        _val(ws1.cell(row=dr, column=1), f"{medal} {rank}", row_fg, bold=(rank == 1))
        _val(ws1.cell(row=dr, column=2), nombre,  row_fg, bold=(rank == 1))
        _val(ws1.cell(row=dr, column=3), email,   row_fg)

        sc = ws1.cell(row=dr, column=4, value=score)
        sc.font  = Font(bold=True, size=11)
        sc.fill  = PatternFill("solid", fgColor=
                    "D1FAE5" if score >= 75 else "FEF9C3" if score >= 50 else "FEE2E2")
        sc.alignment = Alignment(horizontal="center", vertical="center")

        _val(ws1.cell(row=dr, column=5), rec_dis, REC_FILL.get(rec_raw))
        _val(ws1.cell(row=dr, column=6), tech_s,  row_fg)
        _val(ws1.cell(row=dr, column=7), verdict, row_fg)
        ws1.row_dimensions[dr].height = 45

    ws1.column_dimensions["A"].width = 10
    ws1.column_dimensions["B"].width = 28
    ws1.column_dimensions["C"].width = 30
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 16
    ws1.column_dimensions["F"].width = 45
    ws1.column_dimensions["G"].width = 55
    ws1.freeze_panes = "A4"

    # ── Sheet 2: Full detail per candidate ────────────────────────────────────
    ws2 = wb.create_sheet(t["xl_rank_sheet2"])
    r = 1

    for entry in ranked:
        rank    = entry["rank"]
        result  = entry["result"]
        fname   = entry["filename"]
        ctc     = result.get("contacto") or {}
        nombre  = ctc.get("nombre") or fname
        email   = ctc.get("email") or ""
        tel     = ctc.get("telefono") or ""
        li      = ctc.get("linkedin") or ""
        score   = int(result.get("score_general") or 0)
        rec_raw = (result.get("recomendacion") or "").upper()
        rec_dis = t.get(rec_raw, rec_raw)
        habs    = result.get("habilidades") or {}
        puntos  = result.get("puntos_fuertes") or []
        flags   = result.get("red_flags") or []
        verdict = result.get("veredicto_ejecutivo") or ""
        medal   = MEDALS.get(rank, f"#{rank}")

        # Candidate section header
        ws2.merge_cells(f"A{r}:D{r}")
        ch = ws2.cell(row=r, column=1, value=f"{medal} {rank}. {nombre}")
        ch.font  = Font(bold=True, size=12, color="FFFFFF")
        ch.fill  = PatternFill("solid", fgColor=BLUE)
        ch.alignment = Alignment(vertical="center")
        ws2.row_dimensions[r].height = 22
        r += 1

        # Contact + score rows
        for fl, fv in [
            (t["xl_rank_file"], fname),
            (t["xl_email"],    email),
            (t["xl_telefono"], tel),
            ("LinkedIn",       li),
            (t["xl_score"],    f"{score}/100"),
            (t["xl_rec"],      rec_dis),
        ]:
            if fv:
                _lbl(ws2.cell(row=r, column=1), fl)
                cv = ws2.cell(row=r, column=2, value=fv)
                cv.font = Font(size=10)
                cv.alignment = Alignment(wrap_text=True, vertical="center")
                if fl == t["xl_rec"]:
                    cv.fill = PatternFill("solid", fgColor=REC_FILL.get(rec_raw, LIGHT))
                ws2.merge_cells(f"B{r}:D{r}")
                r += 1

        # Verdict
        _lbl(ws2.cell(row=r, column=1), t["xl_verdict"])
        vc = ws2.cell(row=r, column=2, value=verdict)
        vc.font = Font(size=10)
        vc.alignment = Alignment(wrap_text=True, vertical="center")
        ws2.merge_cells(f"B{r}:D{r}")
        ws2.row_dimensions[r].height = 50
        r += 1

        # Skills
        for sk_label, sk_key in [
            (t["xl_tech"], "tecnicas"),
            (t["xl_soft"], "blandas"),
            (t["xl_lang"], "idiomas"),
        ]:
            items = habs.get(sk_key) or []
            if items:
                _lbl(ws2.cell(row=r, column=1), sk_label)
                sv = ws2.cell(row=r, column=2, value="\n".join(items))
                sv.font = Font(size=10)
                sv.alignment = Alignment(wrap_text=True, vertical="top")
                ws2.merge_cells(f"B{r}:D{r}")
                ws2.row_dimensions[r].height = max(15, len(items) * 13)
                r += 1

        if puntos:
            _lbl(ws2.cell(row=r, column=1), t["xl_strong"])
            pv = ws2.cell(row=r, column=2, value="\n".join(f"+ {p}" for p in puntos))
            pv.font = Font(size=10)
            pv.fill = PatternFill("solid", fgColor=GREEN)
            pv.alignment = Alignment(wrap_text=True, vertical="top")
            ws2.merge_cells(f"B{r}:D{r}")
            ws2.row_dimensions[r].height = max(15, len(puntos) * 13)
            r += 1

        if flags:
            _lbl(ws2.cell(row=r, column=1), t["xl_flags"])
            fv2 = ws2.cell(row=r, column=2, value="\n".join(f"! {f}" for f in flags))
            fv2.font = Font(size=10)
            fv2.fill = PatternFill("solid", fgColor=RED)
            fv2.alignment = Alignment(wrap_text=True, vertical="top")
            ws2.merge_cells(f"B{r}:D{r}")
            ws2.row_dimensions[r].height = max(15, len(flags) * 13)
            r += 1

        r += 2  # blank gap between candidates

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 65
    ws2.column_dimensions["C"].width = 5
    ws2.column_dimensions["D"].width = 5

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def recommendation_config(rec: str) -> tuple[str, str]:
    return {
        "CONTRATAR":   ("🟢", "#28a745"),
        "ENTREVISTAR": ("🟡", "#ffc107"),
        "DESCARTAR":   ("🔴", "#dc3545"),
    }.get(rec.upper(), ("⚪", "#6c757d"))


# ── Streamlit App ─────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="CV Analyzer / Analizador de CVs",
    page_icon="📄",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Session state defaults ────────────────────────────────────────────────────
for _k, _v in [
    ("lang", "es"),
    ("authenticated", False),
    ("access_key", ""),
    ("result", None),
    ("pdf_bytes", None),
    ("excel_bytes", None),
    ("profile_img", None),
    ("ranking_results", None),   # list of {"rank", "filename", "result"}
    ("ranking_excel", None),     # bytes
]:
    if _k not in st.session_state:
        st.session_state[_k] = _v

# ── Sidebar: language toggle ──────────────────────────────────────────────────
with st.sidebar:
    current_lang = st.session_state.lang
    other_lang   = "en" if current_lang == "es" else "es"
    toggle_label = TEXTS[current_lang]["lang_toggle"]
    if st.button(f"🌐 {toggle_label}", use_container_width=True):
        st.session_state.lang = other_lang
        st.rerun()

t = TEXTS[st.session_state.lang]  # shorthand for current translations

# ── Auth gate ─────────────────────────────────────────────────────────────────
if not st.session_state.authenticated:
    st.title(t["auth_title"])
    st.markdown(t["auth_subtitle"])
    st.divider()

    with st.form("auth_form"):
        key_input = st.text_input(
            t["auth_label"],
            placeholder=t["auth_placeholder"],
            type="password",
        )
        submitted = st.form_submit_button(t["auth_btn"], use_container_width=True)

    if submitted:
        if not key_input.strip():
            st.error(t["auth_error_empty"])
        else:
            status = validate_key(key_input.strip())
            if status == "ok":
                st.session_state.authenticated = True
                st.session_state.access_key = key_input.strip()
                st.rerun()
            elif status == "expired":
                st.error(t["auth_error_expired"])
            else:
                st.error(t["auth_error_invalid"])
    st.stop()

# ── Main app (only reached when authenticated) ────────────────────────────────

st.title(f"📄 {t['page_title']}")
st.markdown(t["page_subtitle"])
st.divider()

# Show remaining uses in sidebar
with st.sidebar:
    uses_left = key_uses_left(st.session_state.access_key)
    st.caption(f"{'Usos restantes' if st.session_state.lang == 'es' else 'Uses remaining'}: **{uses_left} / {MAX_USES}**")

tab_single, tab_ranking = st.tabs([t["mode_single"], t["mode_ranking"]])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — Individual analysis
# ══════════════════════════════════════════════════════════════════════════════
with tab_single:
    left, right = st.columns([1, 1], gap="large")

    with left:
        st.subheader(t["upload_header"])
        uploaded_file = st.file_uploader(
            t["upload_label"],
            type=["pdf"],
            help=t["upload_help"],
        )
        st.caption(t["upload_caption"])

        st.subheader(t["jd_header"])
        job_description = st.text_area(
            t["jd_label"],
            height=180,
            placeholder=t["jd_placeholder"],
        )

        analyze_btn = st.button(
            t["analyze_btn"],
            type="primary",
            disabled=uploaded_file is None,
            use_container_width=True,
        )
        if not uploaded_file:
            st.caption(t["analyze_caption"])

    # ── Analysis on button click ──────────────────────────────────────────────
    if analyze_btn and uploaded_file:
        pdf_bytes_single = uploaded_file.read()

        with st.spinner(t["spinner_extract"]):
            cv_text     = extract_text_from_pdf(pdf_bytes_single, t)
            profile_img = extract_profile_image(pdf_bytes_single)

        if not cv_text:
            st.error(t["err_no_text"])
        else:
            with st.spinner(t["spinner_analyze"]):
                result = analyze_cv(cv_text, job_description, st.session_state.lang, t)

            if result is not None:
                consume_key(st.session_state.access_key)

                with st.spinner(t["spinner_reports"]):
                    pdf_report   = generate_pdf_report(result, t)
                    excel_report = generate_excel_report(result, t)

                st.session_state.result      = result
                st.session_state.pdf_bytes   = pdf_report
                st.session_state.excel_bytes = excel_report
                st.session_state.profile_img = profile_img

    # ── Right column: results ─────────────────────────────────────────────────
    with right:
        result = st.session_state.result

        if result is None:
            st.markdown(
                f"<br><div style='text-align:center;color:gray;font-size:1.1rem'>"
                f"{t['no_results']}</div>",
                unsafe_allow_html=True,
            )
        else:
            # ── Profile card ──
            contacto = result.get("contacto") or {}
            nombre   = contacto.get("nombre") or "Candidato"
            email    = contacto.get("email") or ""
            telefono = contacto.get("telefono") or ""
            linkedin = contacto.get("linkedin") or ""

            with st.container(border=True):
                photo_col, info_col = st.columns([1, 3])
                with photo_col:
                    img    = st.session_state.profile_img
                    avatar = img if img is not None else make_initials_avatar(nombre)
                    st.image(avatar, width=90)
                with info_col:
                    st.markdown(f"### {nombre}")
                    if email:    st.caption(f"📧 {email}")
                    if telefono: st.caption(f"📱 {telefono}")
                    if linkedin: st.caption(f"🔗 {linkedin}")

            st.divider()

            # ── Score + recommendation ──
            score       = int(result.get("score_general") or 0)
            rec_raw     = (result.get("recomendacion") or "").upper()
            rec_display = t.get(rec_raw, rec_raw)
            emoji, color = recommendation_config(rec_raw)

            c1, c2 = st.columns(2)
            with c1:
                st.metric(t["score_label"], f"{score} / 100")
                st.progress(score / 100)
            with c2:
                st.markdown(
                    f"<div style='font-size:.9rem;color:gray;margin-bottom:4px'>{t['rec_label']}</div>"
                    f"<div style='font-size:1.5rem;font-weight:700;color:{color}'>{emoji} {rec_display}</div>",
                    unsafe_allow_html=True,
                )

            st.divider()

            # ── Veredicto ejecutivo ──
            st.markdown(t["verdict_header"])
            st.info(result.get("veredicto_ejecutivo") or t["verdict_na"])

            # ── Habilidades ──
            st.markdown(t["skills_header"])
            habs = result.get("habilidades") or {}
            sk_tec, sk_bla, sk_idi = st.tabs([t["tab_tech"], t["tab_soft"], t["tab_lang"]])

            with sk_tec:
                items = habs.get("tecnicas") or []
                if items:
                    cols = st.columns(3)
                    for i, s in enumerate(items):
                        cols[i % 3].markdown(f"✅ {s}")
                else:
                    st.write(t["no_tech"])

            with sk_bla:
                items = habs.get("blandas") or []
                for s in items:
                    st.markdown(f"✅ {s}")
                if not items:
                    st.write(t["no_soft"])

            with sk_idi:
                items = habs.get("idiomas") or []
                for s in items:
                    st.markdown(f"🌐 {s}")
                if not items:
                    st.write(t["no_lang"])

            # ── Puntos fuertes & Red flags ──
            col_pos, col_neg = st.columns(2)
            with col_pos:
                st.markdown(t["strong_header"])
                for p in (result.get("puntos_fuertes") or []):
                    st.success(f"✓ {p}")
                if not result.get("puntos_fuertes"):
                    st.write(t["strong_none"])

            with col_neg:
                st.markdown(t["flags_header"])
                for f in (result.get("red_flags") or []):
                    st.warning(f"⚠ {f}")
                if not result.get("red_flags"):
                    st.success(t["flags_none"])

            # ── Download buttons ──
            st.divider()
            st.markdown(t["download_header"])

            slug = (nombre.split()[0] if nombre else "cv").lower()
            dl1, dl2 = st.columns(2)
            with dl1:
                st.download_button(
                    label=t["dl_pdf"],
                    data=st.session_state.pdf_bytes,
                    file_name=f"cv_analysis_{slug}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            with dl2:
                st.download_button(
                    label=t["dl_excel"],
                    data=st.session_state.excel_bytes,
                    file_name=f"cv_analysis_{slug}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

            with st.expander(t["json_expander"]):
                st.json(result)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — Candidate ranking
# ══════════════════════════════════════════════════════════════════════════════
with tab_ranking:
    rank_left, rank_right = st.columns([1, 1], gap="large")

    with rank_left:
        rank_jd = st.text_area(
            t["rank_jd_label"],
            height=160,
            placeholder=t["rank_jd_placeholder"],
            key="rank_jd",
        )

        rank_files = st.file_uploader(
            t["rank_upload_label"],
            type=["pdf"],
            accept_multiple_files=True,
            help=t["rank_upload_help"],
            key="rank_files",
        )
        # Silently cap at 10 files
        if rank_files and len(rank_files) > 10:
            rank_files = rank_files[:10]
            st.caption("⚠ Máximo 10 CVs / Max 10 CVs.")

        st.caption(t["rank_caption"])

        _rank_disabled  = not rank_jd.strip() or not rank_files or len(rank_files) < 2
        _rank_hint = (
            t["rank_btn_need_jd"]    if not rank_jd.strip()
            else t["rank_btn_need_files"] if not rank_files or len(rank_files) < 2
            else ""
        )
        rank_btn = st.button(
            t["rank_btn"],
            type="primary",
            disabled=_rank_disabled,
            use_container_width=True,
        )
        if _rank_hint:
            st.caption(_rank_hint)

    # ── Ranking analysis ──────────────────────────────────────────────────────
    if rank_btn and rank_files and rank_jd.strip():
        collected = []
        failed    = 0
        prog_bar  = st.progress(0, text="")

        for idx, f in enumerate(rank_files, 1):
            prog_bar.progress(
                (idx - 1) / len(rank_files),
                text=t["rank_spinner"].format(i=idx, n=len(rank_files), name=f.name),
            )
            raw_bytes = f.read()
            cv_text   = extract_text_from_pdf(raw_bytes, t)
            if not cv_text:
                failed += 1
                continue

            result, err = analyze_cv_silent(cv_text, rank_jd, st.session_state.lang)
            if result is None:
                failed += 1
                continue

            consume_key(st.session_state.access_key)
            collected.append({"filename": f.name, "result": result})

        prog_bar.progress(1.0, text="✅ Done" if st.session_state.lang == "en" else "✅ Listo")

        if failed:
            st.warning(t["rank_warn_failed"].format(n=failed))

        if len(collected) < 2:
            st.error(t["rank_err_min"])
        else:
            # Sort by score descending, assign rank
            collected.sort(key=lambda x: int(x["result"].get("score_general") or 0), reverse=True)
            for i, entry in enumerate(collected, 1):
                entry["rank"] = i

            ranking_excel = generate_ranking_excel(collected, rank_jd, t)

            st.session_state.ranking_results = collected
            st.session_state.ranking_excel   = ranking_excel

    # ── Ranking results ───────────────────────────────────────────────────────
    with rank_right:
        ranked = st.session_state.ranking_results

        if not ranked:
            st.markdown(
                f"<br><div style='text-align:center;color:gray;font-size:1.1rem'>"
                f"{t['rank_no_results']}</div>",
                unsafe_allow_html=True,
            )
        else:
            MEDALS = {1: "🥇", 2: "🥈", 3: "🥉"}
            REC_COLORS_UI = {
                "CONTRATAR":   "#28a745", "HIRE":      "#28a745",
                "ENTREVISTAR": "#e6a817", "INTERVIEW": "#e6a817",
                "DESCARTAR":   "#dc3545", "DISCARD":   "#dc3545",
            }

            st.markdown(t["rank_results_header"])

            # ── Compact comparison table ──────────────────────────────────────
            tbl_rows = []
            for entry in ranked:
                rank    = entry["rank"]
                result  = entry["result"]
                ctc     = result.get("contacto") or {}
                nombre  = ctc.get("nombre") or entry["filename"]
                score   = int(result.get("score_general") or 0)
                rec_raw = (result.get("recomendacion") or "").upper()
                rec_dis = t.get(rec_raw, rec_raw)
                techs   = (result.get("habilidades") or {}).get("tecnicas") or []
                tbl_rows.append({
                    t["rank_table_pos"]:    f"{MEDALS.get(rank, rank)} {rank}",
                    t["rank_table_name"]:   nombre,
                    t["rank_table_score"]:  score,
                    t["rank_table_rec"]:    rec_dis,
                    t["rank_table_skills"]: ", ".join(techs[:4]) if techs else "—",
                })

            st.dataframe(
                tbl_rows,
                use_container_width=True,
                hide_index=True,
                column_config={
                    t["rank_table_score"]: st.column_config.ProgressColumn(
                        t["rank_table_score"],
                        min_value=0,
                        max_value=100,
                        format="%d",
                    ),
                },
            )

            st.divider()

            # ── Download ranking Excel ────────────────────────────────────────
            st.download_button(
                label=t["rank_download"],
                data=st.session_state.ranking_excel,
                file_name="ranking_candidatos.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )

            st.divider()

            # ── Expandable detail cards per candidate ─────────────────────────
            st.markdown(t["rank_detail_header"])

            for entry in ranked:
                rank    = entry["rank"]
                result  = entry["result"]
                ctc     = result.get("contacto") or {}
                nombre  = ctc.get("nombre") or entry["filename"]
                score   = int(result.get("score_general") or 0)
                rec_raw = (result.get("recomendacion") or "").upper()
                rec_dis = t.get(rec_raw, rec_raw)
                rec_col = REC_COLORS_UI.get(rec_raw, "#6c757d")
                medal   = MEDALS.get(rank, f"#{rank}")
                emoji_r, _ = recommendation_config(rec_raw)

                with st.expander(f"{medal} **{rank}. {nombre}** — {score}/100  {emoji_r} {rec_dis}"):
                    d1, d2 = st.columns([1, 1])

                    with d1:
                        st.metric(t["score_label"], f"{score} / 100")
                        st.progress(score / 100)
                        st.markdown(
                            f"<div style='font-size:.85rem;color:gray'>{t['rec_label']}</div>"
                            f"<div style='font-size:1.2rem;font-weight:700;color:{rec_col}'>"
                            f"{emoji_r} {rec_dis}</div>",
                            unsafe_allow_html=True,
                        )
                        email = ctc.get("email") or ""
                        tel   = ctc.get("telefono") or ""
                        li    = ctc.get("linkedin") or ""
                        if email: st.caption(f"📧 {email}")
                        if tel:   st.caption(f"📱 {tel}")
                        if li:    st.caption(f"🔗 {li}")

                    with d2:
                        st.markdown(f"**{t['verdict_header'].replace('### ', '')}**")
                        st.info(result.get("veredicto_ejecutivo") or t["verdict_na"])

                    habs    = result.get("habilidades") or {}
                    techs   = habs.get("tecnicas") or []
                    soft    = habs.get("blandas") or []
                    langs   = habs.get("idiomas") or []
                    puntos  = result.get("puntos_fuertes") or []
                    flags   = result.get("red_flags") or []

                    if techs:
                        st.markdown(f"**{t['tab_tech']}:** " + " · ".join(f"`{s}`" for s in techs))
                    if soft:
                        st.markdown(f"**{t['tab_soft']}:** " + ", ".join(soft))
                    if langs:
                        st.markdown(f"**{t['tab_lang']}:** " + ", ".join(langs))

                    if puntos or flags:
                        sp_col, fl_col = st.columns(2)
                        with sp_col:
                            if puntos:
                                st.markdown(f"**{t['strong_header'].replace('### ', '')}**")
                                for p in puntos:
                                    st.success(f"✓ {p}")
                        with fl_col:
                            if flags:
                                st.markdown(f"**{t['flags_header'].replace('### ', '')}**")
                                for f in flags:
                                    st.warning(f"⚠ {f}")
